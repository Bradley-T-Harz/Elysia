"""
Bounded local data execution v0 for Elysia.

This module intentionally stays small: CSV support is stdlib-only, and XLSX
support is local/read-only through the optional openpyxl dependency when it is
installed in the active Python environment.

It does not run arbitrary Python, shell commands, web access, file mutation,
plotting, artifact writing, folder scanning, notebook behavior, or memory
promotion.
"""

from __future__ import annotations

from collections.abc import Iterable
import csv
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any


DEFAULT_MAX_FILE_SIZE_BYTES = 1_000_000
DEFAULT_MAX_ROWS = 5_000
DEFAULT_PREVIEW_ROW_LIMIT = 5
MAX_PREVIEW_CELL_LENGTH = 160

DATA_EXECUTOR_TOOL_KIND = "data_executor"
DATA_OPERATION_SUMMARIZE_CSV = "summarize_csv"
SUPPORTED_DATA_FILE_EXTENSIONS = {".csv", ".xlsx"}
MISSING_VALUE_TOKENS = {"", "na", "n/a", "null", "none", "nan", "-"}


class DataExecutionStatus(str, Enum):
    """Small execution status vocabulary for bounded data execution."""

    COMPLETED = "completed"
    FAILED = "failed"
    BLOCKED = "blocked"


@dataclass
class DataExecutionResult:
    """Structured result for a bounded local data execution request."""

    ok: bool
    status: DataExecutionStatus
    tool_kind: str = DATA_EXECUTOR_TOOL_KIND
    operation: str = DATA_OPERATION_SUMMARIZE_CSV
    source_path: str = ""
    file_name: str | None = None
    file_kind: str | None = None
    row_count: int = 0
    column_count: int = 0
    columns: list[str] = field(default_factory=list)
    numeric_columns: list[str] = field(default_factory=list)
    text_columns: list[str] = field(default_factory=list)
    missing_values_by_column: dict[str, int] = field(default_factory=dict)
    preview_rows: list[dict[str, str]] = field(default_factory=list)
    numeric_stats: dict[str, dict[str, float | int | None]] = field(
        default_factory=dict
    )
    locality: str = "local"
    approval_required: bool = False
    network_access_used: bool = False
    mutated_files: bool = False
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _result(
    *,
    ok: bool,
    status: DataExecutionStatus,
    source_path: str | Path,
    file_kind: str | None = None,
    errors: list[str] | None = None,
    warnings: list[str] | None = None,
    **kwargs: Any,
) -> DataExecutionResult:
    path = Path(source_path)
    return DataExecutionResult(
        ok=ok,
        status=status,
        source_path=str(path),
        file_name=path.name if str(path) else None,
        file_kind=file_kind,
        errors=list(errors or []),
        warnings=list(warnings or []),
        **kwargs,
    )


def _file_kind_label(file_kind: str | None) -> str:
    text = str(file_kind or "data").strip()
    return text.upper() if text else "DATA"


def _is_missing(value: str) -> bool:
    return value.strip().lower() in MISSING_VALUE_TOKENS


def _parse_number(value: str) -> float | None:
    text = value.strip()
    if _is_missing(text):
        return None

    try:
        number = float(text)
    except ValueError:
        return None

    if number != number or number in {float("inf"), float("-inf")}:
        return None

    return number


def _truncate_cell(value: str, limit: int = MAX_PREVIEW_CELL_LENGTH) -> str:
    text = str(value)
    if len(text) <= limit:
        return text

    return f"{text[: max(1, limit - 1)].rstrip()}…"


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _validate_header(
    header: Iterable[Any],
    source_path: str | Path,
    *,
    file_kind: str,
) -> DataExecutionResult | None:
    raw_header = list(header)

    if not raw_header:
        return _result(
            ok=False,
            status=DataExecutionStatus.FAILED,
            source_path=source_path,
            file_kind=file_kind,
            errors=[
                f"{_file_kind_label(file_kind)} file is empty or missing a header row."
            ],
        )

    normalized = [_cell_to_text(column).strip() for column in raw_header]
    blank_positions = [
        index + 1 for index, column in enumerate(normalized) if not column
    ]

    if blank_positions:
        return _result(
            ok=False,
            status=DataExecutionStatus.BLOCKED,
            source_path=source_path,
            file_kind=file_kind,
            errors=[
                f"{_file_kind_label(file_kind)} header contains a blank column name. "
                f"Blank header positions: {', '.join(map(str, blank_positions))}."
            ],
        )

    seen: set[str] = set()
    duplicates: list[str] = []

    for column in normalized:
        if column in seen and column not in duplicates:
            duplicates.append(column)
        seen.add(column)

    if duplicates:
        return _result(
            ok=False,
            status=DataExecutionStatus.BLOCKED,
            source_path=source_path,
            file_kind=file_kind,
            errors=[
                f"{_file_kind_label(file_kind)} header contains duplicate column names: "
                + ", ".join(duplicates)
                + "."
            ],
        )

    return None


def _check_local_data_file_boundary(
    path: Path,
    *,
    allowed_suffixes: set[str],
    file_kind: str | None,
    max_file_size_bytes: int,
) -> DataExecutionResult | None:
    if not path.exists():
        return _result(
            ok=False,
            status=DataExecutionStatus.FAILED,
            source_path=path,
            file_kind=file_kind,
            errors=["Data file does not exist."],
        )

    if path.is_dir():
        return _result(
            ok=False,
            status=DataExecutionStatus.BLOCKED,
            source_path=path,
            file_kind=file_kind,
            errors=["Directory paths are not accepted for bounded data execution v0."],
        )

    suffix = path.suffix.lower()
    if suffix not in allowed_suffixes:
        supported = " and ".join(sorted(allowed_suffixes))
        return _result(
            ok=False,
            status=DataExecutionStatus.BLOCKED,
            source_path=path,
            file_kind=file_kind,
            errors=[
                f"Unsupported data file type for v0. Only {supported} are supported."
            ],
        )

    try:
        size_bytes = path.stat().st_size
    except OSError as exc:
        return _result(
            ok=False,
            status=DataExecutionStatus.FAILED,
            source_path=path,
            file_kind=file_kind,
            errors=[f"Could not inspect data file size: {exc}"],
        )

    if size_bytes > max_file_size_bytes:
        return _result(
            ok=False,
            status=DataExecutionStatus.BLOCKED,
            source_path=path,
            file_kind=file_kind,
            errors=[
                f"{_file_kind_label(file_kind)} file size exceeds the bounded data execution v0 limit. "
                f"size_bytes={size_bytes}; max_file_size_bytes={max_file_size_bytes}."
            ],
        )

    return None


def _check_local_csv_boundary(
    path: Path,
    *,
    max_file_size_bytes: int,
) -> DataExecutionResult | None:
    return _check_local_data_file_boundary(
        path,
        allowed_suffixes={".csv"},
        file_kind="csv",
        max_file_size_bytes=max_file_size_bytes,
    )


def _load_openpyxl_load_workbook() -> tuple[Any | None, str | None]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return None, str(exc)

    return load_workbook, None


def _summarize_table_rows(
    *,
    source_path: str | Path,
    file_kind: str,
    header: Iterable[Any],
    rows: Iterable[Iterable[Any]],
    max_rows: int,
    preview_row_limit: int,
    warnings: list[str] | None = None,
) -> DataExecutionResult:
    header_result = _validate_header(header, source_path, file_kind=file_kind)
    if header_result is not None:
        return header_result

    columns = [_cell_to_text(column).strip() for column in header]
    column_count = len(columns)

    missing_values_by_column = {column: 0 for column in columns}
    numeric_values_by_column: dict[str, list[float]] = {
        column: [] for column in columns
    }
    non_missing_counts = {column: 0 for column in columns}
    non_numeric_counts = {column: 0 for column in columns}
    preview_rows: list[dict[str, str]] = []
    final_warnings: list[str] = list(warnings or [])
    row_count = 0
    shorter_row_warning_added = False
    longer_row_warning_added = False
    label = _file_kind_label(file_kind)

    for raw_row in rows:
        row_count += 1

        if row_count > max_rows:
            return _result(
                ok=False,
                status=DataExecutionStatus.BLOCKED,
                source_path=source_path,
                file_kind=file_kind,
                row_count=row_count,
                column_count=column_count,
                columns=columns,
                errors=[
                    f"{label} row count exceeds the bounded data execution v0 limit. "
                    f"max_rows={max_rows}."
                ],
            )

        normalized_row = [_cell_to_text(value) for value in list(raw_row)]

        if len(normalized_row) < column_count and not shorter_row_warning_added:
            final_warnings.append(
                f"One or more {label} rows had fewer cells than the header; missing trailing cells were treated as blank."
            )
            shorter_row_warning_added = True

        if len(normalized_row) > column_count and not longer_row_warning_added:
            final_warnings.append(
                f"One or more {label} rows had more cells than the header; extra cells were ignored for v0 summary."
            )
            longer_row_warning_added = True

        normalized_row = list(normalized_row[:column_count])
        if len(normalized_row) < column_count:
            normalized_row.extend([""] * (column_count - len(normalized_row)))

        if len(preview_rows) < preview_row_limit:
            preview_rows.append(
                {
                    column: _truncate_cell(normalized_row[index])
                    for index, column in enumerate(columns)
                }
            )

        for index, column in enumerate(columns):
            value = normalized_row[index]

            if _is_missing(value):
                missing_values_by_column[column] += 1
                continue

            non_missing_counts[column] += 1
            parsed_number = _parse_number(value)

            if parsed_number is None:
                non_numeric_counts[column] += 1
            else:
                numeric_values_by_column[column].append(parsed_number)

    if row_count == 0:
        final_warnings.append(f"{label} contains a header row but no data rows.")

    numeric_columns = [
        column
        for column in columns
        if non_missing_counts[column] > 0 and non_numeric_counts[column] == 0
    ]
    text_columns = [
        column
        for column in columns
        if row_count > 0 and column not in numeric_columns
    ]

    numeric_stats: dict[str, dict[str, float | int | None]] = {}
    for column in numeric_columns:
        values = numeric_values_by_column[column]
        if not values:
            continue

        numeric_stats[column] = {
            "count": len(values),
            "missing": missing_values_by_column[column],
            "min": min(values),
            "max": max(values),
            "mean": sum(values) / len(values),
        }

    return _result(
        ok=True,
        status=DataExecutionStatus.COMPLETED,
        source_path=source_path,
        file_kind=file_kind,
        row_count=row_count,
        column_count=len(columns),
        columns=columns,
        numeric_columns=numeric_columns,
        text_columns=text_columns,
        missing_values_by_column=missing_values_by_column,
        preview_rows=preview_rows,
        numeric_stats=numeric_stats,
        warnings=final_warnings,
    )


def summarize_csv_file(
    source_path: str | Path,
    *,
    max_file_size_bytes: int = DEFAULT_MAX_FILE_SIZE_BYTES,
    max_rows: int = DEFAULT_MAX_ROWS,
    preview_row_limit: int = DEFAULT_PREVIEW_ROW_LIMIT,
) -> DataExecutionResult:
    """
    Summarize a local CSV file using only bounded stdlib read-only inspection.
    """
    path = Path(source_path)

    boundary_result = _check_local_csv_boundary(
        path,
        max_file_size_bytes=max_file_size_bytes,
    )
    if boundary_result is not None:
        return boundary_result

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)

            try:
                raw_header = next(reader)
            except StopIteration:
                return _result(
                    ok=False,
                    status=DataExecutionStatus.FAILED,
                    source_path=path,
                    file_kind="csv",
                    errors=["CSV file is empty or missing a header row."],
                )

            return _summarize_table_rows(
                source_path=path,
                file_kind="csv",
                header=raw_header,
                rows=reader,
                max_rows=max_rows,
                preview_row_limit=preview_row_limit,
            )

    except UnicodeDecodeError:
        return _result(
            ok=False,
            status=DataExecutionStatus.FAILED,
            source_path=path,
            file_kind="csv",
            errors=["CSV file could not be decoded as UTF-8 text."],
        )
    except csv.Error as exc:
        return _result(
            ok=False,
            status=DataExecutionStatus.FAILED,
            source_path=path,
            file_kind="csv",
            errors=[f"CSV parser failed: {exc}"],
        )
    except OSError as exc:
        return _result(
            ok=False,
            status=DataExecutionStatus.FAILED,
            source_path=path,
            file_kind="csv",
            errors=[f"CSV file could not be read: {exc}"],
        )


def summarize_xlsx_file(
    source_path: str | Path,
    *,
    max_file_size_bytes: int = DEFAULT_MAX_FILE_SIZE_BYTES,
    max_rows: int = DEFAULT_MAX_ROWS,
    preview_row_limit: int = DEFAULT_PREVIEW_ROW_LIMIT,
) -> DataExecutionResult:
    """
    Summarize a local XLSX workbook using bounded read-only openpyxl inspection.

    Sprint 5C v0 inspects only the first worksheet. Elysia does not evaluate
    spreadsheet formulas; with openpyxl data_only=True it reads cached cell
    values when present.
    """
    path = Path(source_path)

    boundary_result = _check_local_data_file_boundary(
        path,
        allowed_suffixes={".xlsx"},
        file_kind="xlsx",
        max_file_size_bytes=max_file_size_bytes,
    )
    if boundary_result is not None:
        return boundary_result

    load_workbook, import_error = _load_openpyxl_load_workbook()
    if load_workbook is None:
        return _result(
            ok=False,
            status=DataExecutionStatus.BLOCKED,
            source_path=path,
            file_kind="xlsx",
            errors=[
                "XLSX support requires the optional local dependency openpyxl. "
                "CSV support remains available without it. "
                f"Import error: {import_error}"
            ],
        )

    workbook = None
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
        )

        sheet_names = list(getattr(workbook, "sheetnames", []) or [])
        if not sheet_names:
            return _result(
                ok=False,
                status=DataExecutionStatus.FAILED,
                source_path=path,
                file_kind="xlsx",
                errors=["XLSX workbook does not contain any worksheets."],
            )

        worksheet = workbook[sheet_names[0]]
        rows = worksheet.iter_rows(values_only=True)

        try:
            raw_header = next(rows)
        except StopIteration:
            return _result(
                ok=False,
                status=DataExecutionStatus.FAILED,
                source_path=path,
                file_kind="xlsx",
                errors=["XLSX file is empty or missing a header row."],
            )

        warnings = [
            "XLSX v0 reads cached cell values only; formulas are not evaluated by Elysia."
        ]
        if len(sheet_names) > 1:
            warnings.append(
                f"XLSX v0 inspected only the first worksheet: {sheet_names[0]}."
            )

        return _summarize_table_rows(
            source_path=path,
            file_kind="xlsx",
            header=raw_header,
            rows=rows,
            max_rows=max_rows,
            preview_row_limit=preview_row_limit,
            warnings=warnings,
        )

    except Exception as exc:
        return _result(
            ok=False,
            status=DataExecutionStatus.FAILED,
            source_path=path,
            file_kind="xlsx",
            errors=[f"XLSX workbook could not be read: {exc}"],
        )
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def summarize_data_file(
    source_path: str | Path,
    *,
    max_file_size_bytes: int = DEFAULT_MAX_FILE_SIZE_BYTES,
    max_rows: int = DEFAULT_MAX_ROWS,
    preview_row_limit: int = DEFAULT_PREVIEW_ROW_LIMIT,
) -> DataExecutionResult:
    """
    Dispatch bounded data execution by file type.

    v0 supports CSV and optional local XLSX table summaries only.
    """
    path = Path(source_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return summarize_csv_file(
            path,
            max_file_size_bytes=max_file_size_bytes,
            max_rows=max_rows,
            preview_row_limit=preview_row_limit,
        )

    if suffix == ".xlsx":
        return summarize_xlsx_file(
            path,
            max_file_size_bytes=max_file_size_bytes,
            max_rows=max_rows,
            preview_row_limit=preview_row_limit,
        )

    return _result(
        ok=False,
        status=DataExecutionStatus.BLOCKED,
        source_path=path,
        errors=["Unsupported data file type for v0. Only .csv and .xlsx are supported."],
    )


__all__ = (
    "DATA_EXECUTOR_TOOL_KIND",
    "DATA_OPERATION_SUMMARIZE_CSV",
    "DEFAULT_MAX_FILE_SIZE_BYTES",
    "DEFAULT_MAX_ROWS",
    "DEFAULT_PREVIEW_ROW_LIMIT",
    "DataExecutionResult",
    "DataExecutionStatus",
    "SUPPORTED_DATA_FILE_EXTENSIONS",
    "summarize_csv_file",
    "summarize_data_file",
    "summarize_xlsx_file",
)

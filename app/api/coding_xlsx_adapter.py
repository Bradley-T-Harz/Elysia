"""XLSX adapter for bounded workbook inspection; formulas are inert text."""

from __future__ import annotations

from pathlib import Path
from typing import Any


def _cell_value(value: object) -> str:
    if value is None:
        return ""
    return str(value)[:500]


def extract_xlsx_preview(path: Path, *, max_chars: int, max_tables: int, max_rows: int) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=False, data_only=False)
    metadata = {
        "sheet_count": len(workbook.worksheets),
        "sheet_names": [sheet.title for sheet in workbook.worksheets],
        "creator": workbook.properties.creator,
        "title": workbook.properties.title,
        "created": workbook.properties.created.isoformat() if workbook.properties.created else None,
        "modified": workbook.properties.modified.isoformat() if workbook.properties.modified else None,
    }
    tables: list[dict[str, Any]] = []
    outline: list[dict[str, Any]] = []
    provenance: list[dict[str, Any]] = []
    text_parts: list[str] = []
    formula_count = 0

    for sheet_index, sheet in enumerate(workbook.worksheets[:max_tables], start=1):
        rows: list[list[str]] = []
        for row in sheet.iter_rows(min_row=1, max_row=min(max_rows, sheet.max_row), values_only=False):
            rendered_row: list[str] = []
            for cell in row[: min(sheet.max_column, 12)]:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    rendered_row.append("[formula redacted as inert text]")
                else:
                    rendered_row.append(_cell_value(value))
            rows.append(rendered_row)
        tables.append(
            {
                "kind": "sheet_preview",
                "sheet": sheet.title,
                "sheet_index": sheet_index,
                "rows": rows,
                "row_count_previewed": len(rows),
                "column_count_previewed": max((len(row) for row in rows), default=0),
                "dimensions": sheet.calculate_dimension(),
                "hidden": sheet.sheet_state != "visible",
            }
        )
        outline.append({"kind": "sheet", "sheet": sheet.title, "dimensions": sheet.calculate_dimension(), "hidden": sheet.sheet_state != "visible"})
        provenance.append({"kind": "sheet", "sheet": sheet.title, "rows_previewed": len(rows)})
        if len("\n".join(text_parts)) < max_chars:
            text_parts.append(f"[sheet {sheet.title}]\n" + "\n".join("\t".join(row) for row in rows[:max_rows]))

    metadata["formula_count_previewed"] = formula_count
    return {
        "status": "completed",
        "metadata": metadata,
        "text_preview": "\n\n".join(text_parts)[:max_chars],
        "tables": tables,
        "outline": outline,
        "provenance": provenance,
        "warnings": ["Formula cells are reported as inert metadata and are never executed."] if formula_count else [],
    }


__all__ = ("extract_xlsx_preview",)
